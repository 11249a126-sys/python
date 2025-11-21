import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "DemoSheet"
sheet['A1'] = "Name"
sheet['B1'] = "Age"
sheet['A2'] = "Alice"
sheet['B2'] = 25
sheet['A3'] = "Bob"
sheet['B3'] = 30
wb.save("demo.xlsx")
print("Data written to demo.xlsx successfully!")
wb = openpyxl.load_workbook("demo.xlsx")
sheet = wb["DemoSheet"]
for row in sheet.iter_rows(values_only=True):
    print(row)
